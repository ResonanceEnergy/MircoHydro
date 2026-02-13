#!/usr/bin/env python3
"""MicroHydroV1 - Add Run_ID column + backfill migrator for Measurements sheet.

What it does
- Opens an Excel workbook (default: automation/MicroHydroV1_RnD_Export.xlsx)
- Ensures the Measurements sheet has a Run_ID column header (adds it if missing)
- Backfills Run_ID for rows where it is blank by extracting the run id token from:
  - Conditions (default)
  - Notes
  - Sensor (optional)

Default run id patterns supported
- 2026-01-22_RUN3
- 2026-01-22_Run3
- 2026-01-22_Run3_FR1 (extracts the prefix 2026-01-22_RUN3)
- 2026-01-22-RUN3 (also supported)

Safety
- Creates a timestamped backup copy of the workbook before writing changes.

Usage
  python tools/workflow/migrate_add_run_id.py --workbook automation/MicroHydroV1_RnD_Export.xlsx

Write to a new output file instead of in-place
  python tools/workflow/migrate_add_run_id.py --workbook automation/MicroHydroV1_RnD_Export.xlsx     --out automation/MicroHydroV1_RnD_Export_with_RunID.xlsx

If you use a different Run_ID format
  python tools/workflow/migrate_add_run_id.py --workbook ... --pattern "(YOUR_REGEX_HERE)"

"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Optional, Dict


def default_pattern() -> re.Pattern:
    # Capture run id token like 2026-01-22_RUN3 or 2026-01-22_Run3 or 2026-01-22-RUN3
    # Allows optional suffixes after the token; we capture only the token.
    return re.compile(r'(\d{4}-\d{2}-\d{2}[_-](?:RUN|Run)\d+)', re.IGNORECASE)


def extract_run_id(text: str, pat: re.Pattern) -> Optional[str]:
    if not text:
        return None
    m = pat.search(str(text))
    if not m:
        return None
    token = m.group(1)
    # Normalize to YYYY-MM-DD_RUNN format (RUN uppercase)
    token = re.sub(r'[_-](?:run)', '_RUN', token, flags=re.IGNORECASE)
    token = token.replace('-RUN', '_RUN')
    return token


def load_headers(ws) -> Dict[str, int]:
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[str(v).strip()] = c
    return headers


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('--workbook', required=True, help='Path to MicroHydroV1_RnD_Export.xlsx')
    ap.add_argument('--out', default=None, help='Optional output workbook path (default: in-place)')
    ap.add_argument('--sheet', default='Measurements', help='Sheet name (default: Measurements)')
    ap.add_argument('--pattern', default=None, help='Regex pattern with one capture group for Run_ID token')
    ap.add_argument('--fields', nargs='*', default=['Conditions', 'Notes'], help='Columns to scan for Run_ID')
    args = ap.parse_args(argv)

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        raise SystemExit('Workbook not found: ' + str(wb_path))

    out_path = Path(args.out) if args.out else wb_path

    # Backup (always)
    ts = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = wb_path.with_name(wb_path.stem + f'_BACKUP_{ts}' + wb_path.suffix)
    backup_path.write_bytes(wb_path.read_bytes())

    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' not found in workbook")
    ws = wb[args.sheet]

    headers = load_headers(ws)

    # Ensure Run_ID column
    if 'Run_ID' in headers:
        run_col = headers['Run_ID']
    else:
        run_col = ws.max_column + 1
        ws.cell(1, run_col).value = 'Run_ID'
        headers['Run_ID'] = run_col

    # Pattern
    pat = re.compile(args.pattern, re.IGNORECASE) if args.pattern else default_pattern()

    # Backfill
    total = 0
    filled = 0
    already = 0
    missing = 0

    # Pre-resolve field columns
    field_cols = []
    for f in args.fields:
        if f in headers:
            field_cols.append((f, headers[f]))
    # If none found, still run but won't fill

    for r in range(2, ws.max_row + 1):
        total += 1
        cur = ws.cell(r, run_col).value
        if cur and str(cur).strip():
            already += 1
            continue

        found = None
        for name, col in field_cols:
            txt = ws.cell(r, col).value
            found = extract_run_id(str(txt) if txt is not None else '', pat)
            if found:
                break

        if found:
            ws.cell(r, run_col).value = found
            filled += 1
        else:
            missing += 1

    wb.save(out_path)

    print('[OK] Workbook saved:', out_path)
    print('[OK] Backup created:', backup_path)
    print('[STATS] rows:', total, 'already:', already, 'filled:', filled, 'missing:', missing)

    if missing:
        print('[NOTE] Some rows have no Run_ID. Ensure Conditions/Notes contain tokens like 2026-01-22_RUN3.')


if __name__ == '__main__':
    main()
