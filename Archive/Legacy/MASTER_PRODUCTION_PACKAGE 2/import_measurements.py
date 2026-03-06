#!/usr/bin/env python3
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXCEL_NAME = 'MicroHydroV1_RnD_Export.xlsx'
SHEET_NAME = 'Measurements'
TABLE_NAME = 'tblMeasurements'

REQUIRED = ['Timestamp','Test_ID','Sensor','Value','Unit']
OPTIONAL = ['Baseline_Value','Conditions','Notes']
ALL_COLS = REQUIRED + OPTIONAL

PASS_SR = (
    '=IF(AND([@Test_ID]="T-001",ISNUMBER(SEARCH("coherence",LOWER([@Sensor]))),N([@Value])>=0.9),TRUE,'
    'IF(AND([@Test_ID]="T-002",ISNUMBER(SEARCH("ripple",LOWER([@Sensor]))),N([@Baseline_Value])>0,N([@Value])<=N([@Baseline_Value])*0.70),TRUE,'
    'IF(OR([@Test_ID]="",[@Sensor]="",[@Value]=""),"",FALSE)))'
)
RULE_SR = (
    '=IF([@Test_ID]="T-001","Value>=0.90 (coherence)",IF([@Test_ID]="T-002","Value<=Baseline*0.70 (≥30% reduction)",""))'
)

def read_csv_any(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    # Normalize columns: case-insensitive matching to expected names
    map_cols = {c.lower(): c for c in df.columns}
    out = {}
    for col in ALL_COLS:
        # pick matching source col by case-insensitive
        src = None
        for k, v in map_cols.items():
            if k == col.lower():
                src = v; break
        out[col] = df[src] if src in df.columns else None
    new_df = pd.DataFrame(out)
    # Ensure required present
    missing = [c for c in REQUIRED if new_df[c] is None or new_df[c].isna().all()]
    if missing:
        raise ValueError(f"Missing required columns in {path.name}: {missing}")
    # Fill optional with '' if missing
    for c in OPTIONAL:
        if new_df[c] is None:
            new_df[c] = ''
    return new_df


def import_csvs(files: list[str]) -> dict:
    # discover if none provided
    if not files:
        files = [p.name for p in Path('.').glob('*.csv') if p.name != 'Measurements_Import_Template.csv']
    if not files:
        return {"imported": 0, "by_test": {}, "files": []}

    frames = []
    for fn in files:
        p = Path(fn)
        if not p.exists():
            continue
        df = read_csv_any(p)
        df['Source_File'] = p.name
        frames.append(df)

    if not frames:
        return {"imported": 0, "by_test": {}, "files": []}

    all_rows = pd.concat(frames, ignore_index=True)

    # Open workbook
    wb = load_workbook(EXCEL_NAME)
    ws = wb[SHEET_NAME]

    # Find table bounds
    tbl = None
    for t in ws._tables:
        if t.name == TABLE_NAME:
            tbl = t; break
    if tbl is None:
        raise RuntimeError(f"Table {TABLE_NAME} not found in {SHEET_NAME}")

    # Determine insert start row (end of table + 1)
    from openpyxl.utils.cell import coordinate_to_tuple
    ref = tbl.ref  # e.g., 'A1:J2'
    start_cell, end_cell = ref.split(':')
    _, end_row = coordinate_to_tuple(end_cell)
    insert_row = end_row + 1

    # Write rows
    by_test = {}
    for _, row in all_rows.iterrows():
        values = [row.get(c, '') if pd.notna(row.get(c, '')) else '' for c in ALL_COLS]
        # Columns A..H are data, I..J are formulas
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=insert_row, column=col_idx).value = val
        ws.cell(row=insert_row, column=9).value = PASS_SR
        ws.cell(row=insert_row, column=10).value = RULE_SR
        # Next row
        tid = str(row.get('Test_ID','')).strip()
        by_test[tid] = by_test.get(tid, 0) + 1
        insert_row += 1

    # Expand table ref to new last row
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(ws.max_column if ws.max_column>=10 else 10)
    tbl.ref = f"A1:{last_col_letter}{insert_row-1}"

    wb.save(EXCEL_NAME)

    return {"imported": len(all_rows), "by_test": by_test, "files": files}


if __name__ == '__main__':
    files = sys.argv[1:]
    result = import_csvs(files)
    print(result)
