#!/usr/bin/env python3
"""MicroHydroV1 - Measurements importer (patched to stamp Run_ID).

Adds:
- Run_ID column support in Measurements sheet
- Automatic Run_ID stamping per imported row

Run_ID behavior
- If CSV already includes a Run_ID column, use it.
- Else, if --run-id is provided, stamp that for all rows.
- Else, attempt to derive Run_ID from file path (folder or filename) using token like YYYY-MM-DD_RUN3.
- Else leave blank.

Usage:
  python automation/import_measurements.py <csv1> <csv2> ...

Optional:
  --run-id 2026-01-22_RUN3
  --excel automation/MicroHydroV1_RnD_Export.xlsx

"""

import sys
import re
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

EXCEL_NAME_DEFAULT = 'MicroHydroV1_RnD_Export.xlsx'
SHEET_NAME = 'Measurements'
TABLE_NAME = 'tblMeasurements'

REQUIRED = ['Timestamp','Test_ID','Sensor','Value','Unit']
OPTIONAL = ['Baseline_Value','Conditions','Notes','Run_ID']
ALL_COLS = REQUIRED + OPTIONAL

PASS_SR = (
    '=IF(AND([@Test_ID]="T-001",ISNUMBER(SEARCH("coherence",LOWER([@Sensor]))),N([@Value])>=0.9),TRUE,'
    'IF(AND([@Test_ID]="T-002",ISNUMBER(SEARCH("ripple",LOWER([@Sensor]))),N([@Baseline_Value])>0,N([@Value])<=N([@Baseline_Value])*0.70),TRUE,'
    'IF(OR([@Test_ID]="",[@Sensor]="",[@Value]=""),"",FALSE)))'
)
RULE_SR = (
    '=IF([@Test_ID]="T-001","Value>=0.90 (coherence)",IF([@Test_ID]="T-002","Value<=Baseline*0.70 (≥30% reduction)",""))'
)

RUNID_RE = re.compile(r'(\d{4}-\d{2}-\d{2}[_-](?:RUN|Run)\d+)', re.IGNORECASE)


def normalize_run_id(token: str) -> str:
    token = token.strip()
    # normalize separators and RUN casing
    token = token.replace('-RUN', '_RUN')
    token = re.sub(r'[_-](?:run)', '_RUN', token, flags=re.IGNORECASE)
    return token


def derive_run_id_from_path(path: Path) -> str:
    s = str(path)
    m = RUNID_RE.search(s)
    if m:
        return normalize_run_id(m.group(1))
    return ''


def read_csv_any(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    # Normalize columns: case-insensitive matching to expected names
    map_cols = {c.lower(): c for c in df.columns}
    out = {}
    for col in ALL_COLS:
        src = map_cols.get(col.lower())
        if src is not None:
            out[col] = df[src]
        else:
            out[col] = ''
    return pd.DataFrame(out)


def ensure_run_id_column(ws, tbl=None):
    """Ensure the worksheet has a Run_ID column header. Adds it at the end if missing."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[str(v).strip()] = c
    if 'Run_ID' in headers:
        return headers

    # append new column at the end
    new_col = ws.max_column + 1
    ws.cell(1, new_col).value = 'Run_ID'
    headers['Run_ID'] = new_col

    # expand table ref if available
    if tbl is not None:
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(ws.max_column)
        tbl.ref = f"A1:{last_col_letter}{ws.max_row}"

    return headers


def import_csvs(files, excel_path: Path, run_id: str = ''):
    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME]

    # locate the table
    tbl = None
    if hasattr(ws, 'tables'):
        tbl = ws.tables.get(TABLE_NAME)

    headers = ensure_run_id_column(ws, tbl)

    # verify required columns exist in worksheet
    for col in REQUIRED + ['Pass','Rule_Eval']:
        if col not in headers:
            raise SystemExit(f"Workbook Measurements missing required column header: {col}")

    all_rows = []
    for f in files:
        p = Path(f)
        df = read_csv_any(p)
        # stamp run_id for this file if not already in CSV
        rid = run_id or derive_run_id_from_path(p)
        if (df['Run_ID'] == '').all():
            df['Run_ID'] = rid
        all_rows.append(df)

    if not all_rows:
        return {"imported": 0, "by_test": {}, "files": []}

    big = pd.concat(all_rows, ignore_index=True)

    # Filter: must have required fields
    for col in REQUIRED:
        if col not in big.columns:
            raise SystemExit(f"Missing column in import data: {col}")

    # Prepare insertion point
    insert_row = ws.max_row + 1

    by_test = {}

    pass_col = headers['Pass']
    rule_col = headers['Rule_Eval']

    # write rows
    for _, row in big.iterrows():
        # write mapped columns
        for col in REQUIRED + OPTIONAL:
            col_idx = headers.get(col)
            if col_idx:
                ws.cell(row=insert_row, column=col_idx).value = row.get(col, '')

        # formulas
        ws.cell(row=insert_row, column=pass_col).value = PASS_SR
        ws.cell(row=insert_row, column=rule_col).value = RULE_SR

        tid = str(row.get('Test_ID','')).strip()
        by_test[tid] = by_test.get(tid, 0) + 1
        insert_row += 1

    # Expand table ref to new last row
    if tbl is not None:
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(ws.max_column)
        tbl.ref = f"A1:{last_col_letter}{insert_row-1}"

    wb.save(excel_path)

    return {"imported": int(len(big)), "by_test": by_test, "files": list(map(str, files)), "run_id": run_id}


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('csv', nargs='+', help='CSV files to import')
    ap.add_argument('--run-id', default='', help='Optional Run_ID to stamp (e.g., 2026-01-22_RUN3)')
    ap.add_argument('--excel', default=EXCEL_NAME_DEFAULT, help='Path to MicroHydroV1_RnD_Export.xlsx')
    args = ap.parse_args(argv)

    excel_path = Path(args.excel)
    if not excel_path.exists():
        # allow running from same folder as workbook
        excel_path = Path(EXCEL_NAME_DEFAULT)
    if not excel_path.exists():
        raise SystemExit('Excel workbook not found. Provide --excel path.')

    result = import_csvs(args.csv, excel_path=excel_path, run_id=args.run_id)
    print(result)


if __name__ == '__main__':
    main()
